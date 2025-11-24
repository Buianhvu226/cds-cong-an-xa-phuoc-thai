from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class ExportService:
    def __init__(self):
        self.export_dir = 'exports'
        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir)
        self.type_names = {
            'weapons': 'Vũ khí, VLN, CCHT',
            'vehicles': 'Phương tiện',
            'water': 'Trang thiết bị thủy',
            'technical': 'Thiết bị KTNV',
            'office': 'Thiết bị VP & DT'
        }
    
    def export_assets_to_excel(self, assets, asset_type, filters=None):
        """Export assets to Excel file"""
        wb = Workbook()
        ws = wb.active
        self.write_asset_sheet(ws, asset_type, assets)
        
        # Save file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{asset_type}_{timestamp}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def write_asset_sheet(self, ws, asset_type, assets, sheet_title=None):
        """Populate a worksheet with asset data"""
        title = sheet_title or f"Danh sách {self.type_names.get(asset_type, asset_type)}"
        ws.title = title[:31]
        
        headers = self.get_headers(asset_type)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row_idx, asset in enumerate(assets, start=2):
            values = self.get_asset_values(asset, asset_type)
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20
    
    def get_headers(self, asset_type):
        """Get headers based on asset type"""
        if asset_type == 'weapons':
            return [
                'Mã tài sản', 'Danh mục', 'Tên tài sản', 'Đơn vị tính',
                'Số lượng', 'Nguyên giá', 'Giá trị còn lại', 'Số hiệu',
                'Loại tài sản', 'Thực tế bàn giao', 'Định kỳ kiểm tra',
                'Ngày kiểm tra gần nhất', 'Ngày kiểm tra tiếp theo',
                'Kết quả kiểm tra', 'Năm hết hạn', 'Ghi chú'
            ]
        elif asset_type == 'vehicles':
            return [
                'Mã tài sản', 'Danh mục PT', 'Tên phương tiện', 'Đơn vị tính',
                'Nguyên giá', 'Số lượng', 'Biển số', 'Số khung/Số thân vỏ',
                'Số máy', 'Năm trang bị', 'Loại tài sản', 'Thực tế bàn giao',
                'Ngày đăng kiểm', 'Ngày thay nhớt', 'Ngày thay vỏ',
                'Định kỳ kiểm tra', 'Ngày kiểm tra gần nhất', 'Ngày kiểm tra tiếp theo',
                'Kết quả kiểm tra', 'Ghi chú'
            ]
        elif asset_type == 'technical':
            return [
                'Mã tài sản', 'Tên thiết bị', 'Năm sử dụng', 'Số lượng',
                'Nguyên giá', 'Giá trị còn lại', 'Loại tài sản', 'Thực tế bàn giao',
                'Định kỳ kiểm tra', 'Ngày kiểm tra gần nhất', 'Ngày kiểm tra tiếp theo',
                'Kết quả kiểm tra', 'Năm hết hạn', 'Ghi chú'
            ]
        elif asset_type == 'office':
            return [
                'Mã tài sản', 'Tên thiết bị', 'Năm sử dụng', 'Số lượng',
                'Nguyên giá', 'Giá trị còn lại', 'Loại tài sản', 'Thực tế bàn giao',
                'Hình thức', 'Sự kiện', 'Chi phí', 'Định kỳ kiểm tra',
                'Ngày kiểm tra gần nhất', 'Ngày kiểm tra tiếp theo',
                'Kết quả kiểm tra', 'Năm hết hạn', 'Ghi chú'
            ]
        else:  # default
            return [
                'Mã tài sản', 'Tên tài sản', 'Đơn vị tính', 'Số lượng',
                'Nguyên giá', 'Giá trị còn lại', 'Loại tài sản', 'Ghi chú'
            ]
    
    def get_asset_values(self, asset, asset_type):
        """Get values for asset row"""
        if asset_type == 'weapons':
            return [
                asset.get('ma_tai_san', ''),
                asset.get('ma_danh_muc', ''),
                asset.get('ten_tai_san', ''),
                asset.get('don_vi_tinh', ''),
                asset.get('so_luong', ''),
                asset.get('nguyen_gia', ''),
                asset.get('gia_tri_con_lai', ''),
                asset.get('so_hieu', ''),
                asset.get('loai_tai_san', ''),
                asset.get('thuc_te_ban_giao', ''),
                asset.get('dinh_ky_kiem_tra', ''),
                asset.get('ngay_kiem_tra_gan_nhat', ''),
                asset.get('ngay_kiem_tra_tiep_theo', ''),
                asset.get('ket_qua_kiem_tra', ''),
                asset.get('nam_het_han', ''),
                asset.get('ghi_chu', '')
            ]
        elif asset_type == 'vehicles':
            return [
                asset.get('ma_tai_san', ''),
                asset.get('danh_muc_phuong_tien', ''),
                asset.get('ten_phuong_tien', ''),
                asset.get('don_vi_tinh', ''),
                asset.get('nguyen_gia', ''),
                asset.get('so_luong', ''),
                asset.get('bien_so_ky_hieu', ''),
                asset.get('so_khung_so_than_vo', ''),
                asset.get('so_may', ''),
                asset.get('nam_trang_bi', ''),
                asset.get('loai_tai_san', ''),
                asset.get('thuc_te_ban_giao', ''),
                asset.get('ngay_dang_kiem', ''),
                asset.get('ngay_thay_nhot', ''),
                asset.get('ngay_thay_vo', ''),
                asset.get('dinh_ky_kiem_tra', ''),
                asset.get('ngay_kiem_tra_gan_nhat', ''),
                asset.get('ngay_kiem_tra_tiep_theo', ''),
                asset.get('ket_qua_kiem_tra', ''),
                asset.get('ghi_chu', '')
            ]
        elif asset_type == 'technical':
            return [
                asset.get('ma_tai_san', ''),
                asset.get('ten_tai_san', ''),
                asset.get('nam_su_dung', ''),
                asset.get('so_luong', ''),
                asset.get('nguyen_gia', ''),
                asset.get('gia_tri_con_lai', ''),
                asset.get('loai_tai_san', ''),
                asset.get('thuc_te_ban_giao', ''),
                asset.get('dinh_ky_kiem_tra', ''),
                asset.get('ngay_kiem_tra_gan_nhat', ''),
                asset.get('ngay_kiem_tra_tiep_theo', ''),
                asset.get('ket_qua_kiem_tra', ''),
                asset.get('nam_het_han', ''),
                asset.get('ghi_chu', '')
            ]
        else:  # office
            return [
                asset.get('ma_tai_san', ''),
                asset.get('ten_tai_san', ''),
                asset.get('nam_su_dung', ''),
                asset.get('so_luong', ''),
                asset.get('nguyen_gia', ''),
                asset.get('gia_tri_con_lai', ''),
                asset.get('loai_tai_san', ''),
                asset.get('thuc_te_ban_giao', ''),
                asset.get('hinh_thuc', ''),
                asset.get('su_kien', ''),
                asset.get('chi_phi', ''),
                asset.get('dinh_ky_kiem_tra', ''),
                asset.get('ngay_kiem_tra_gan_nhat', ''),
                asset.get('ngay_kiem_tra_tiep_theo', ''),
                asset.get('ket_qua_kiem_tra', ''),
                asset.get('nam_het_han', ''),
                asset.get('ghi_chu', '')
            ]

