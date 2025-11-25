from models import (
    DanhSachVuKhiCongCuHoTro,
    DanhSachPhuongTien,
    DanhSachThietBiKyThuatNghiepVu,
    DanhSachThietBiVanPhongDoanhTrai,
    DanhSachTrangThietBiThuy
)
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, Reference
from services.export_service import ExportService
import os
from sqlalchemy import or_, and_

class ReportService:
    def __init__(self):
        self.models = {
            'weapons': DanhSachVuKhiCongCuHoTro,
            'vehicles': DanhSachPhuongTien,
            'water': DanhSachTrangThietBiThuy,
            'technical': DanhSachThietBiKyThuatNghiepVu,
            'office': DanhSachThietBiVanPhongDoanhTrai
        }
        self.type_order = ['weapons', 'vehicles', 'water', 'technical', 'office']
        self.type_names = {
            'weapons': 'Vũ khí, VLN, CCHT',
            'vehicles': 'Phương tiện',
            'water': 'Trang thiết bị thủy',
            'technical': 'Thiết bị KTNV',
            'office': 'Thiết bị VP & DT'
        }
    
    def generate_report(self, report_type, filters=None):
        """Generate report by type"""
        if report_type == 'summary':
            return self._generate_summary_report()
        elif report_type == 'by_category':
            category = filters.get('category', 'all') if filters else 'all'
            return self._generate_category_report(category)
        elif report_type == 'inspection_due':
            days = int(filters.get('days', 30)) if filters else 30
            return self._generate_inspection_due_report(days)
        elif report_type == 'expiring':
            days = int(filters.get('days', 30)) if filters else 30
            return self._generate_expiring_report(days)
        else:
            raise ValueError(f"Invalid report type: {report_type}")
    
    def _generate_summary_report(self):
        """Generate summary report"""
        # Similar to dashboard stats but more detailed
        return {'type': 'summary', 'data': []}
    
    def _generate_category_report(self, category):
        """Generate report by category"""
        return {'type': 'by_category', 'category': category, 'data': []}
    
    def _generate_inspection_due_report(self, days=30):
        """Generate report for assets due for inspection in next N days"""
        today = date.today()
        end_date = today + timedelta(days=days)
        
        all_assets = []
        
        for asset_type, model_class in self.models.items():
            assets = model_class.query.filter(
                model_class.is_deleted == False,
                model_class.ngay_kiem_tra_tiep_theo.isnot(None),
                model_class.ngay_kiem_tra_tiep_theo >= today,
                model_class.ngay_kiem_tra_tiep_theo <= end_date
            ).all()
            
            for asset in assets:
                asset_dict = asset.to_dict()
                asset_dict['asset_type'] = asset_type
                all_assets.append(asset_dict)
        
        # Sort by date
        all_assets.sort(key=lambda x: x.get('ngay_kiem_tra_tiep_theo', ''))
        
        return {'type': 'inspection_due', 'days': days, 'data': all_assets}
    
    def _generate_expiring_report(self, days=30):
        """
        Generate report for expiring assets.
        Theo yêu cầu nghiệp vụ: hiển thị toàn bộ tài sản trong hệ thống (kể cả chưa tới hạn),
        đảm bảo người dùng nhìn tổng thể rồi dùng filter phụ (frontend) để đánh dấu cái sắp hết hạn.
        """
        today = date.today()
        end_date = today + timedelta(days=days)

        assets_by_type = self._collect_assets_by_type()
        totals_by_type = {
            asset_type: len(assets_by_type.get(asset_type, []))
            for asset_type in self.type_order
        }

        all_assets = []
        for asset_type in self.type_order:
            for asset in assets_by_type.get(asset_type, []):
                asset_copy = dict(asset)
                asset_copy['asset_type'] = asset_type
                all_assets.append(asset_copy)

        default_future = end_date + timedelta(days=3650)
        all_assets.sort(
            key=lambda x: (
                self._parse_date(x.get('ngay_kiem_tra_tiep_theo')) or default_future,
                x.get('nam_het_han') or 9999
            )
        )

        return {
            'type': 'expiring',
            'days': days,
            'total_assets': sum(totals_by_type.values()),
            'by_type_totals': totals_by_type,
            'data': all_assets
        }
    
    def export_report(self, report_type, filters=None):
        """Export report to Excel"""
        export_service = ExportService()
        assets_by_type = self._collect_assets_by_type()
        summary = self._build_summary_data(assets_by_type)
        
        wb = Workbook()
        summary_ws = wb.active
        self._build_summary_sheet(summary_ws, summary)
        
        # Create detail sheets for each asset type
        for asset_type in self.type_order:
            assets = assets_by_type.get(asset_type, [])
            ws = wb.create_sheet()
            sheet_title = self.type_names.get(asset_type, asset_type)
            export_service.write_asset_sheet(ws, asset_type, assets, sheet_title=sheet_title)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"bao_cao_{report_type}_{timestamp}.xlsx"
        filepath = os.path.join(export_service.export_dir, filename)
        wb.save(filepath)
        return filepath
    
    def _collect_assets_by_type(self):
        assets_by_type = {}
        for asset_type, model_class in self.models.items():
            assets = model_class.query.filter_by(is_deleted=False).all()
            assets_by_type[asset_type] = [asset.to_dict() for asset in assets]
        return assets_by_type
    
    def _build_summary_data(self, assets_by_type):
        today = date.today()
        due_threshold = today + timedelta(days=15)
        
        summary = {
            'generated_at': datetime.now(),
            'total_assets': 0,
            'counts_by_type': {},
            'value_by_type': {},
            'status_counts': {'overdue': 0, 'due_soon': 0, 'normal': 0},
            'total_nguyen_gia': 0,
            'total_gia_tri_con_lai': 0
        }
        
        for asset_type in self.type_order:
            assets = assets_by_type.get(asset_type, [])
            summary['counts_by_type'][asset_type] = len(assets)
            summary['total_assets'] += len(assets)
            
            ng_sum = 0
            remain_sum = 0
            
            for asset in assets:
                ng_sum += self._to_number(asset.get('nguyen_gia'))
                remain_sum += self._to_number(asset.get('gia_tri_con_lai'))
                
                next_date = self._parse_date(asset.get('ngay_kiem_tra_tiep_theo'))
                if next_date:
                    if next_date < today:
                        summary['status_counts']['overdue'] += 1
                    elif next_date <= due_threshold:
                        summary['status_counts']['due_soon'] += 1
                    else:
                        summary['status_counts']['normal'] += 1
            
            summary['value_by_type'][asset_type] = {
                'nguyen_gia': ng_sum,
                'gia_tri_con_lai': remain_sum
            }
            summary['total_nguyen_gia'] += ng_sum
            summary['total_gia_tri_con_lai'] += remain_sum
        
        return summary
    
    def _build_summary_sheet(self, ws, summary):
        ws.title = 'Tổng hợp'
        ws.merge_cells('A1:F1')
        ws['A1'] = 'BÁO CÁO & THỐNG KÊ TÀI SẢN'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = 'Ngày xuất'
        ws['B2'] = summary['generated_at'].strftime('%d/%m/%Y %H:%M')
        ws['A2'].font = Font(bold=True)
        ws['B2'].alignment = Alignment(horizontal='left')
        
        ws['A3'] = 'Tổng số tài sản'
        ws['B3'] = summary['total_assets']
        ws['A3'].font = Font(bold=True)
        ws['B3'].number_format = '#,##0'
        
        ws['A4'] = 'Tổng nguyên giá (VND)'
        ws['B4'] = summary['total_nguyen_gia']
        ws['A4'].font = Font(bold=True)
        ws['B4'].number_format = '#,##0'
        
        ws['A5'] = 'Tổng giá trị còn lại (VND)'
        ws['B5'] = summary['total_gia_tri_con_lai']
        ws['A5'].font = Font(bold=True)
        ws['B5'].number_format = '#,##0'
        
        ws['D2'] = 'Quá hạn'
        ws['E2'] = summary['status_counts']['overdue']
        ws['D3'] = 'Sắp hết hạn (≤15 ngày)'
        ws['E3'] = summary['status_counts']['due_soon']
        ws['D4'] = 'Còn hạn'
        ws['E4'] = summary['status_counts']['normal']
        for cell in ['D2', 'D3', 'D4']:
            ws[cell].font = Font(bold=True)
        for cell in ['E2', 'E3', 'E4']:
            ws[cell].number_format = '#,##0'
        
        type_table_start = 7
        headers = ['Danh mục', 'Số lượng', 'Tổng nguyên giá (VND)', 'Giá trị còn lại (VND)']
        header_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=type_table_start, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        current_row = type_table_start + 1
        for asset_type in self.type_order:
            row_values = [
                self.type_names.get(asset_type, asset_type),
                summary['counts_by_type'].get(asset_type, 0),
                summary['value_by_type'].get(asset_type, {}).get('nguyen_gia', 0),
                summary['value_by_type'].get(asset_type, {}).get('gia_tri_con_lai', 0)
            ]
            for col, value in enumerate(row_values, start=1):
                cell = ws.cell(row=current_row, column=col, value=value)
                if col == 1:
                    cell.alignment = Alignment(horizontal='left')
                else:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
            current_row += 1
        
        # Pie chart for distribution by category
        chart = PieChart()
        data_ref = Reference(ws, min_col=2, min_row=type_table_start, max_row=current_row - 1)
        labels_ref = Reference(ws, min_col=1, min_row=type_table_start + 1, max_row=current_row - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        chart.title = "Phân bổ tài sản"
        chart.height = 7
        chart.width = 12
        ws.add_chart(chart, f'G{type_table_start}')
        
        ws.column_dimensions['A'].width = 26
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['G'].width = 2
    
    def _to_number(self, value):
        if value is None:
            return 0
        try:
            return float(value)
        except (TypeError, ValueError):
            try:
                return float(str(value).replace(',', ''))
            except Exception:
                return 0
    
    def _parse_date(self, value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            try:
                return datetime.fromisoformat(value).date()
            except ValueError:
                return None
        return None
