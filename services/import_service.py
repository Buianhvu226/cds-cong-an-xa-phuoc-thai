from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from dateutil import parser as date_parser

class ImportService:
    def __init__(self):
        self.import_dir = 'imports'
        if not os.path.exists(self.import_dir):
            os.makedirs(self.import_dir)
    
    def create_template(self, asset_type):
        """Create Excel template for import"""
        wb = Workbook()
        ws = wb.active
        
        # Set title
        type_names = {
            'weapons': 'Vũ khí, VLN, CCHT',
            'vehicles': 'Phương tiện',
            'water': 'Trang thiết bị thủy',
            'technical': 'Thiết bị KTNV',
            'office': 'Thiết bị VP & DT'
        }
        title = f"Mẫu nhập {type_names.get(asset_type, asset_type)}"
        ws.title = title[:31]
        
        # Get headers and field mappings
        headers, field_mappings, required_fields = self._get_field_config(asset_type)
        
        # Header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add example row with instructions
        example_row = 2
        ws.cell(row=example_row, column=1, value="VÍ DỤ:")
        example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=example_row, column=col)
            cell.fill = example_fill
            cell.font = Font(italic=True)
        
        # Add example data based on asset type
        example_data = self._get_example_data(asset_type)
        for col, value in enumerate(example_data, start=1):
            ws.cell(row=example_row, column=col, value=value)
        
        # Add instruction row
        instruction_row = 3
        ws.merge_cells(f'A{instruction_row}:{get_column_letter(len(headers))}{instruction_row}')
        instruction_cell = ws.cell(row=instruction_row, column=1, value="HƯỚNG DẪN: Xóa dòng ví dụ và nhập dữ liệu của bạn. Các cột có dấu * là bắt buộc.")
        instruction_cell.font = Font(bold=True, color="FF0000")
        instruction_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 25
        
        # Save file
        filename = f"mau_nhap_{asset_type}.xlsx"
        filepath = os.path.join(self.import_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def import_from_excel(self, filepath, asset_type):
        """Import assets from Excel file - generator that yields row data"""
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Get field configuration
            headers, field_mappings, required_fields = self._get_field_config(asset_type)
            
            # Validate headers
            excel_headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    excel_headers.append(str(header).strip())
            
            # Check if headers match
            if len(excel_headers) != len(headers):
                yield {'success': False, 'errors': [f"Số lượng cột không khớp. Yêu cầu: {len(headers)}, Tìm thấy: {len(excel_headers)}"], 'success_count': 0, 'skipped_count': 0}
                return
            
            # Map Excel columns to field names
            column_mapping = {}
            header_errors = []
            for idx, excel_header in enumerate(excel_headers):
                # Try to find matching header (case-insensitive, with/without asterisk)
                excel_header_clean = excel_header.replace('*', '').strip()
                found = False
                for header_idx, expected_header in enumerate(headers):
                    expected_clean = expected_header.replace('*', '').strip()
                    if excel_header_clean.lower() == expected_clean.lower():
                        column_mapping[idx] = field_mappings[header_idx]
                        found = True
                        break
                if not found:
                    header_errors.append(f"Cột '{excel_header}' không khớp với cấu trúc mẫu")
            
            if header_errors:
                yield {'success': False, 'errors': header_errors, 'success_count': 0, 'skipped_count': 0}
                return
            
            # Process data rows (skip header and example rows)
            start_row = 2
            for row_idx in range(start_row, ws.max_row + 1):
                row_data = {}
                
                # Skip empty rows
                if all(ws.cell(row=row_idx, column=col).value is None for col in range(1, ws.max_column + 1)):
                    continue
                
                # Extract data from row
                for col_idx, field_name in column_mapping.items():
                    cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
                    
                    # Convert value based on field type
                    converted_value = self._convert_value(cell_value, field_name, asset_type)
                    if converted_value is not None:
                        row_data[field_name] = converted_value
                
                # Validate required fields
                missing_fields = []
                for field in required_fields:
                    if field not in row_data or row_data[field] is None or row_data[field] == '':
                        missing_fields.append(field)
                
                if missing_fields:
                    # Skip this row, will be counted as error by caller
                    continue
                
                # Validate data format
                validation_errors = self._validate_row_data(row_data, asset_type, row_idx)
                if validation_errors:
                    # Skip this row, will be counted as error by caller
                    continue
                
                # Row is valid, yield it for processing
                yield row_data
                
        except Exception as e:
            yield {'success': False, 'errors': [f"Lỗi khi đọc file Excel: {str(e)}"], 'success_count': 0, 'skipped_count': 0}
    
    def _get_field_config(self, asset_type):
        """Get field configuration for asset type"""
        if asset_type == 'weapons':
            headers = [
                'Mã tài sản', 'Danh mục*', 'Tên tài sản*', 'Đơn vị tính*',
                'Số lượng*', 'Nguyên giá', 'Giá trị còn lại', 'Số hiệu',
                'Năm sử dụng', 'Loại tài sản', 'Thực tế bàn giao',
                'Định kỳ kiểm tra', 'Ngày kiểm tra gần nhất', 'Ngày kiểm tra tiếp theo',
                'Kết quả kiểm tra', 'Năm hết hạn', 'Phương thức xử lý', 'Ghi chú'
            ]
            field_mappings = [
                'ma_tai_san', 'ma_danh_muc', 'ten_tai_san', 'don_vi_tinh',
                'so_luong', 'nguyen_gia', 'gia_tri_con_lai', 'so_hieu',
                'nam_su_dung', 'loai_tai_san', 'thuc_te_ban_giao',
                'dinh_ky_kiem_tra', 'ngay_kiem_tra_gan_nhat', 'ngay_kiem_tra_tiep_theo',
                'ket_qua_kiem_tra', 'nam_het_han', 'phuong_thuc_xu_ly', 'ghi_chu'
            ]
            required_fields = ['ma_danh_muc', 'ten_tai_san', 'don_vi_tinh', 'so_luong']
        elif asset_type == 'vehicles':
            headers = [
                'Mã tài sản', 'Danh mục PT*', 'Tên phương tiện*', 'Đơn vị tính*',
                'Nguyên giá', 'Số lượng*', 'Biển số', 'Số khung/Số thân vỏ',
                'Số máy', 'Năm trang bị', 'Loại tài sản', 'Thực tế bàn giao',
                'Ngày đăng kiểm', 'Ngày thay nhớt', 'Ngày thay vỏ',
                'Phí đường bộ', 'Định kỳ kiểm tra', 'Ngày kiểm tra gần nhất',
                'Ngày kiểm tra tiếp theo', 'Kết quả kiểm tra', 'Năm hết hạn',
                'Phương thức xử lý', 'Ghi chú'
            ]
            field_mappings = [
                'ma_tai_san', 'danh_muc_phuong_tien', 'ten_phuong_tien', 'don_vi_tinh',
                'nguyen_gia', 'so_luong', 'bien_so_ky_hieu', 'so_khung_so_than_vo',
                'so_may', 'nam_trang_bi', 'loai_tai_san', 'thuc_te_ban_giao',
                'ngay_dang_kiem', 'ngay_thay_nhot', 'ngay_thay_vo',
                'phi_duong_bo', 'dinh_ky_kiem_tra', 'ngay_kiem_tra_gan_nhat',
                'ngay_kiem_tra_tiep_theo', 'ket_qua_kiem_tra', 'nam_het_han',
                'phuong_thuc_xu_ly', 'ghi_chu'
            ]
            required_fields = ['danh_muc_phuong_tien', 'ten_phuong_tien', 'don_vi_tinh', 'so_luong']
        elif asset_type == 'water':
            headers = [
                'Mã tài sản', 'Danh mục*', 'Tên trang bị*', 'Đơn vị tính*',
                'Nguyên giá', 'Số lượng*', 'Mã hiệu', 'Năm trang bị',
                'Loại tài sản', 'Định kỳ kiểm tra chất lượng', 'Kết quả kiểm tra',
                'Năm hết hạn', 'Phương thức xử lý', 'Ghi chú'
            ]
            field_mappings = [
                'ma_tai_san', 'danh_muc_trang_thiet_bi', 'ten_trang_bi', 'don_vi_tinh',
                'nguyen_gia', 'so_luong', 'ma_hieu', 'nam_trang_bi',
                'loai_tai_san', 'ngay_kiem_tra_gan_nhat', 'ket_qua_kiem_tra',
                'nam_het_han', 'phuong_thuc_xu_ly', 'ghi_chu'
            ]
            required_fields = ['danh_muc_trang_thiet_bi', 'ten_trang_bi', 'don_vi_tinh', 'so_luong']
        elif asset_type == 'technical':
            headers = [
                'Mã tài sản', 'Tên đơn vị, tài sản*', 'Năm sử dụng', 'Số lượng*',
                'Nguyên giá', 'Giá trị còn lại', 'Loại tài sản', 'Thực tế bàn giao',
                'Định kỳ kiểm tra chất lượng', 'Kết quả kiểm tra', 'Năm hết hạn',
                'Phương thức xử lý', 'Ghi chú'
            ]
            field_mappings = [
                'ma_tai_san', 'ten_tai_san', 'nam_su_dung', 'so_luong',
                'nguyen_gia', 'gia_tri_con_lai', 'loai_tai_san', 'thuc_te_ban_giao',
                'ngay_kiem_tra_gan_nhat', 'ket_qua_kiem_tra', 'nam_het_han',
                'phuong_thuc_xu_ly', 'ghi_chu'
            ]
            required_fields = ['ten_tai_san', 'so_luong']
        else:  # office
            headers = [
                'Mã tài sản', 'Tên đơn vị, tài sản*', 'Năm sử dụng', 'Số lượng*',
                'Nguyên giá', 'Giá trị còn lại', 'Loại tài sản', 'Thực tế bàn giao',
                'Hình thức', 'Sự kiện', 'Phí', 'Định kỳ kiểm tra chất lượng',
                'Kết quả kiểm tra', 'Năm hết hạn', 'Phương thức xử lý', 'Ghi chú'
            ]
            field_mappings = [
                'ma_tai_san', 'ten_tai_san', 'nam_su_dung', 'so_luong',
                'nguyen_gia', 'gia_tri_con_lai', 'loai_tai_san', 'thuc_te_ban_giao',
                'hinh_thuc', 'su_kien', 'chi_phi', 'ngay_kiem_tra_gan_nhat',
                'ket_qua_kiem_tra', 'nam_het_han', 'phuong_thuc_xu_ly', 'ghi_chu'
            ]
            required_fields = ['ten_tai_san', 'so_luong']
        
        return headers, field_mappings, required_fields
    
    def _get_example_data(self, asset_type):
        """Get example data for template"""
        if asset_type == 'weapons':
            return ['VK2511001', 'Súng', 'Súng AKM', 'Khẩu', 1, 9481024, 9481024, '141759', 2016, 'Tài sản đặc biệt', 'Có', '6 tháng', '2025-06-01', '2025-12-01', 'Đạt', 2028, 'Tiếp tục sử dụng', '']
        elif asset_type == 'vehicles':
            return ['PT2511001', 'Ô tô', 'Suzuki Carry', 'Chiếc', 378800000, 1, '60A-009.87', 'MHYHDC61TMJ912829', 'K15BT1297661', 2024, 'Chuyên dụng', 'Có', '2025-10-01', '2025-08-15', '2025-07-01', 500000, '3 tháng', '2025-08-15', '2025-11-15', 'Đạt', 2028, 'Tiếp tục sử dụng', '']
        elif asset_type == 'water':
            return ['TT2511001', 'Phao áo', 'Phao cứu sinh tiêu chuẩn', 'Chiếc', 500000, 10, 'PA-001', 2024, 'Chuyên dụng', '2025-06-01', 'Đạt', 2028, 'Tiếp tục sử dụng', '']
        elif asset_type == 'technical':
            return ['TB2511001', 'Máy tính xách tay', 2023, 1, 15000000, 12000000, 'Chuyên dụng', 'Có', '2025-06-01', 'Đạt', 2029, 'Tiếp tục sử dụng', '']
        else:  # office
            return ['VP2511001', 'Bàn làm việc', 2023, 5, 5000000, 4000000, 'Quản lý', 'Có', 'Mua mới', 'Sửa chữa', 500000, '2025-06-01', 'Đạt', 2029, 'Tiếp tục sử dụng', '']
    
    def _convert_value(self, value, field_name, asset_type):
        """Convert Excel cell value to appropriate Python type"""
        if value is None:
            return None
        
        # Date fields
        if 'ngay' in field_name or 'date' in field_name:
            if isinstance(value, datetime):
                return value.date()
            elif isinstance(value, str):
                try:
                    return date_parser.parse(value).date()
                except:
                    return None
            return None
        
        # Number fields
        if field_name in ['so_luong', 'nguyen_gia', 'gia_tri_con_lai', 'nam_su_dung', 
                          'nam_trang_bi', 'nam_het_han', 'chi_phi', 'phi_duong_bo']:
            if isinstance(value, (int, float)):
                return int(value) if field_name in ['so_luong', 'nam_su_dung', 'nam_trang_bi', 'nam_het_han'] else float(value)
            elif isinstance(value, str):
                try:
                    return int(value) if field_name in ['so_luong', 'nam_su_dung', 'nam_trang_bi', 'nam_het_han'] else float(value)
                except:
                    return None
            return None
        
        # String fields
        return str(value).strip() if value else None
    
    def _validate_row_data(self, row_data, asset_type, row_num):
        """Validate row data format"""
        errors = []
        
        # Validate numeric fields
        numeric_fields = ['so_luong', 'nguyen_gia', 'gia_tri_con_lai', 'nam_su_dung', 
                         'nam_trang_bi', 'nam_het_han', 'chi_phi', 'phi_duong_bo']
        for field in numeric_fields:
            if field in row_data and row_data[field] is not None:
                if not isinstance(row_data[field], (int, float)):
                    errors.append(f"Dòng {row_num}: '{field}' phải là số")
        
        # Validate date fields
        from datetime import date
        date_fields = [f for f in row_data.keys() if 'ngay' in f]
        for field in date_fields:
            if row_data[field] and not isinstance(row_data[field], (date, datetime, type(None))):
                errors.append(f"Dòng {row_num}: '{field}' không đúng định dạng ngày (dd/mm/yyyy)")
        
        # Validate select fields
        if asset_type == 'weapons':
            if 'ma_danh_muc' in row_data and row_data['ma_danh_muc']:
                valid_options = ['Súng', 'Đạn', 'Công cụ hỗ trợ']
                if row_data['ma_danh_muc'] not in valid_options:
                    errors.append(f"Dòng {row_num}: 'Danh mục' phải là một trong: {', '.join(valid_options)}")
        
        return errors

